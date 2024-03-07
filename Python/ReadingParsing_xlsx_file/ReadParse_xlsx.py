import os
import shutil
from openpyxl import load_workbook
from datetime import datetime
import time
import glob
import xlsxwriter
cwd = os.getcwd()
def main():
    ############## Finding current directory path###############################
    cwd = os.getcwd()

    ############# Finding and loading Dav file in the current directory########
    # davfile = glob.glob(cwd+"/*DAV*.xlsx")
    davfile = glob.glob(cwd+"/*DAV*.xlsx")
    Load_Davfile = load_workbook(davfile[0])

    ############# Loading Dav Report and identify number of rows and column#########
    # DavSheet = Load_Davfile.get_sheet_by_name('DAV')
    DavSheet = Load_Davfile['DAV']
    col = DavSheet.max_column
    row = col = DavSheet.max_row

    ########## Find a Column No. where is 'ConfigTime' #############
    for i in range(1,col):
        if DavSheet.cell(1,i).value == 'configTime':
            col_configtime = i
            break

    ##############Days Difference Testing from today to Dav ConfigTime#############################
    # d1 = DavSheet.cell(2,col_configtime).value
    # d2 = datetime.now()
    # print(d1,d2)
    # numberofdays = abs(d2-d1).days
    # print(numberofdays)
    ####################### Extract List of devices with ConfigTime more than 21 days in Dav Report ####
    stale_dev = []
    d1 = datetime.now()
    for i in range (2,row):
        d2 = DavSheet.cell(i,col_configtime).value
        if d2 == 'no data':
            stale_dev.append(DavSheet.cell(i,1).value)
        else:
            numberofdays = abs(d1-d2).days
            if numberofdays >= 21:
                stale_dev.append(DavSheet.cell(i,1).value)

    ################ List of the CC_files in the given directory##############################
    cc_folder = time.strftime("%Y-%m-%d")
    CC_Files = glob.glob(cc_folder+"//*9K*.xlsx") + glob.glob(cc_folder+"//*1K*.xlsx") \
            + glob.glob(cc_folder+"//*920*.xlsx") + glob.glob(cc_folder+"//*903*.xlsx") \
                + glob.glob(cc_folder+"//*540*.xlsx")+ glob.glob(cc_folder+"//*55*.xlsx")

    ############Read all the CC_Files and Identify Rows matching stale data and delete/copy them.###
    rr = 1
    stale_dev_info = []
    for file in CC_Files:
        load_file = load_workbook(file)
        CC_Sheet = load_file['Sheet1']
        # print(file)
        row = CC_Sheet.max_row

        ## Identify the rows matching to the stale list of devices##
        Stale_Row = []
        for i in range(2,row+1):
            for staledevice in stale_dev:
                if CC_Sheet.cell(i,2).value == staledevice:
                    Stale_Row.append(i)
                    stale_dev_info.append(CC_Sheet.cell(i,1).value + ' ' + CC_Sheet.cell(i,2).value)

        ## Deleting Rows which matched to stale_row##########
        Stale_Row_Reverse = Stale_Row[::-1]
        for i in Stale_Row_Reverse:
            if i == row:
                CC_Sheet.cell(i,1).value = None
                CC_Sheet.cell(i,2).value = None
                CC_Sheet.cell(i,3).value = None
                CC_Sheet.cell(i,4).value = None
            else:
                # print(CC_Sheet.cell(i,2).value)
                # print(i)
                CC_Sheet.delete_rows(i,1)

        load_file.save(file)
        load_file.close()


    ##Stale List written in text file##########
    '''
    StaleDevice = open(('DeviceList_with_StaleConfig_'+ time.strftime("%m-%d-%y") + '.txt'),'w')
    for device in stale_dev_info:
        StaleDevice.write(device + '\n')
    StaleDevice.close()
    '''
    ####Create or load a xlsx file 'DeviceList_withStaleConfig_Date ##########
    wb_filename = cc_folder + '//DeviceList_with_StaleConfig_'+ time.strftime("%m-%d-%y") + '.xlsx'
    StaleConfigFile = glob.glob(cc_folder + "//*StaleConfig*.xlsx")
    if not StaleConfigFile:##Create a File and write the data
        wb_filename = cc_folder + '//DeviceList_with_StaleConfig_'+ str(time.strftime("%m-%d-%y")) + '.xlsx'
        workbook = xlsxwriter.Workbook(wb_filename)
        worksheet = workbook.add_worksheet()
        worksheet.write(0,0, 'Device IP')
        worksheet.write(0,1, 'Device Name')
        stalelist = set(stale_dev_info)
        rr = 1
        #######Stale List written in XLSX#############
        for device in stalelist:
            dev = device.split( )
            worksheet.write(rr, 0, dev[0])
            worksheet.write(rr, 1, dev[1])
            rr = rr + 1
    else: # Read the existing file and append the data
        workbook =  load_workbook(StaleConfigFile[0])
        worksheet = workbook.get_sheet_by_name('Sheet1')
        stalelist = set(stale_dev_info)
        rr = worksheet.max_row + 1
        #######Stale List written in XLSX#############
        for device in stalelist:
            dev = device.split( )
            worksheet.cell(rr, 1).value = dev[0]
            worksheet.cell(rr, 2).value = dev[1]
            rr = rr + 1
        workbook.save(StaleConfigFile[0])

    workbook.close()
    ##############################################
    ######## Delete Dav File from the #################
    try:
        checktodelete = input("\nReady to delete Dav file...Y or N? " )
        if (checktodelete.lower() == 'y' or checktodelete.lower() == 'yes'):
        	os.remove(davfile)
    except Exception as e:
        print(e)



if __name__ == "__main__":
    main()
