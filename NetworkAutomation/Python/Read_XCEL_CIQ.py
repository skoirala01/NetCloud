'''
#############################################################################
### This script was created for use with the Verizon Wireless EBH team.   ###
### Authors:                                                              ###
### Sudip Koirala (skoirala@cisco)                                        ###
###                                                                       ###
#############################################################################
'''

import re, sys, os, json, logging, io, copy, math
from datetime import datetime
from openpyxl import load_workbook                  # Used to read the Excel CIQ file
from mimir import Mimir, MimirAuthenticationError   # Mimir is used to pull running configs from Network Profiler.
from ciscoconfparse import CiscoConfParse           # CiscoConfParse is used to parse the running configs.
import networkx as nx                   # Used for building the topology connections
import matplotlib                       # Used for plotting the topology
import matplotlib.pyplot as plt         # Used for displaying the topology

wb = load_workbook('BLTNMN-20240122-034926.xlsx')
# Specify the sheets that should be parsed during info gathering.
req_sheets = 'Validation Result'

ws = wb[req_sheets]
# Record the headers actually configured in the sheet from the CIQ.
h_list = [x[0].value for x in ws.iter_cols(min_col=None, max_col=ws.max_column, min_row=6, max_row=6)]
print(h_list)

ciq_db = {}
for row in ws.iter_rows():
    if row[h_list.index('Device Name')].row == 7:
        ciq_db[row[h_list.index('Device Name')].value] = [row[h_list.index('Remediation')].value]
        yy = row[h_list.index('Device Name')].value
        print(yy)
    if row[h_list.index('Device Name')].row > 7:
        #print(row[h_list.index('Device Name')].row)
        #print(yy)
        if row[h_list.index('Device Name')].value == yy:
            xx = row[h_list.index('Remediation')].value
            #print(xx)
           # print(yy)
            yy = row[h_list.index('Device Name')].value
            if xx is not None:
                ciq_db[row[h_list.index('Device Name')].value].append(xx)
        else:
            ciq_db[row[h_list.index('Device Name')].value] = [row[h_list.index('Remediation')].value]
            yy = row[h_list.index('Device Name')].value


print(ciq_db)
runfile = CiscoConfParse(ciq_db)
print(runfile)

