# -*- coding: utf-8 -*-
"""
Created on Tue Oct 23 16:21:15 2018

@author: skoirala
"""

from openpyxl import Workbook
from openpyxl import load_workbook


wb1 = load_workbook('abc33.xlsx')

#print wb1.sheetnames

#active_sheet = wb1.active
#print(type(active_sheet))

#wb1.active


sheet = wb1.get_sheet_by_name('Sheet2')

print(sheet.title)
a1 = sheet['A1']
print a1.value
sheet.delete_cols(5,9)
a1 = sheet['A5']
print a1.value
#source =  wb1.copy_worksheet(sheet)

wb1.create_sheet('sid1',0)
ws = wb1.get_sheet_by_name("sid1")
ws['A1'] = sheet['B1'].value



row = sheet.max_row
col = sheet.max_column

for i in range (1,row+1):
    for j in range (1,col+1):
        ws.cell(i,j).value = sheet.cell(i,j).value


ws['B1'] = '150000'
print ws['B1'].value
ws.cell(2,2).value = 2
wb1.save('abc33.xlsx')

#from xlutils.copy import copy as xl_copy
#wb = xl_copy('Sudip')
#wb1.create_sheet('sid1')
#sheet0 = wb1.get_sheet_by_name("sid1")
#sheet0['A1'] = 1500
#
#wb1.save('abc334.xlsx')




#
#wb2 = Workbook()
#
### get Sheet
##source=wb1.get_sheet_by_name('Sudip')
#source = wb1.active
### copy sheet
#target=wb2.copy_worksheet(source)
#wb2
#wb2.save('abc22.xlsx')
#


#wb1.create_sheet("NewData")
#wb1.save('ReadExcel1.xlsx')

#wb1.create_sheet('Sample Sheet')
#wb1.save('Sample1.xlsx')

#wb1.remove_sheet("Sheet2")
#wb1.create_sheet("January", 0)
#print(wb1.sheetnames)
#
#wb1.save('sample.xlsx')
////////////////////////////////////////////////////
# Import `os`
import os

# Retrieve current working directory (`cwd`)
cwd = os.getcwd()
cwd

# Change directory
os.chdir("/path/to/your/folder")

# List all files and directories in current directory
os.listdir('.')



##EXCEL
import xlwt
import xlrd
import random
from os import path

DavReport = xlrd.open_workbook("C:\Users\x\x\x\DavReport.xlsx")
AllData = DavReport.sheet_by_name("xxxx")

for row_index in xrange(1, AllData.nrows):
	##read it

##Writing new sheet
book = xlwt.Workbook(encoding = "utf-8")
sh = boot.add_sheet("IT Rechabele")

sh.write(0,0, data)


======
#sentdex.com/TutSheet.csv
import numpy as np

def main():
try:
 date, rate, arb = np.loadtxt('filename.csv', delimiter = ',', unpack = True, dtype = 'str')

print date[5], rate[5], arb[5]

except Exception, e:
print str(e)

===========
workbook = xlrd.open_workbook(file_location)
sheet = workbook.sheet_by_index(0)
sheet.cell_value(0,0)
sheet.nrows
sheet.ncols

for col in range (sheet.ncols):
 print sheet.cell_value(0,col)

for col in range (sheet.ncols):
 print sheet.cell_value(1,col)

data = [sheet.cell_value(r,c) for c in range(sheet.ncols) for r in range(sheet.nrows)]

type(data) => list
data[0][0]
etc.


##################
most important... www.lexicon.net/sjmachin/xlrd.html
##########3that gives what is that meant.
sheet.cell_type(1,2)
if it returns 3 then it is float.
###########################################





# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="/home/ubuntu/demo.xlsx"
# load demo.xlsx
wb=load_workbook(filepath)
# get Sheet
source=wb.get_sheet_by_name('Sheet')
# copy sheet
target=wb.copy_worksheet(source)
# save workbook
wb.save(filepath)


# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="/home/ubuntu/demo.xlsx"
# load demo.xlsx
wb=load_workbook(filepath)
# create new sheet
wb.remove(wb.get_sheet_by_name('Sheet 2'))
# save workbook
wb.save(filepath)

/////////////////////////////////////////////
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 23 16:21:15 2018

@author: skoirala
"""

from openpyxl import Workbook
from openpyxl import load_workbook


wb1 = load_workbook('abc33.xlsx')

#print wb1.sheetnames

#active_sheet = wb1.active
#print(type(active_sheet))

#wb1.active


sheet = wb1.get_sheet_by_name('Sheet2')

print(sheet.title)
a1 = sheet['A1']
print a1.value
sheet.delete_cols(5,9)
a1 = sheet['A5']
print a1.value
#source =  wb1.copy_worksheet(sheet)

wb1.create_sheet('sid1',0)
ws = wb1.get_sheet_by_name("sid1")
ws['A1'] = sheet['B1'].value



row = sheet.max_row
col = sheet.max_column

for i in range (1,row+1):
    for j in range (1,col+1):
        ws.cell(i,j).value = sheet.cell(i,j).value


ws['B1'] = '150000'
print ws['B1'].value
ws.cell(2,2).value = 2
wb1.save('abc33.xlsx')

#from xlutils.copy import copy as xl_copy
#wb = xl_copy('Sudip')
#wb1.create_sheet('sid1')
#sheet0 = wb1.get_sheet_by_name("sid1")
#sheet0['A1'] = 1500
#
#wb1.save('abc334.xlsx')




#
#wb2 = Workbook()
#
### get Sheet
##source=wb1.get_sheet_by_name('Sudip')
#source = wb1.active
### copy sheet
#target=wb2.copy_worksheet(source)
#wb2
#wb2.save('abc22.xlsx')
#


#wb1.create_sheet("NewData")
#wb1.save('ReadExcel1.xlsx')

#wb1.create_sheet('Sample Sheet')
#wb1.save('Sample1.xlsx')

#wb1.remove_sheet("Sheet2")
#wb1.create_sheet("January", 0)
#print(wb1.sheetnames)
#
#wb1.save('sample.xlsx')
