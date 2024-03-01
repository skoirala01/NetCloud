#!/usr/bin/env python3
from urllib.request import Request, urlopen
import json
import requests
from datetime import datetime
import time
import getpass
import openpyxl
import os
from lxml import etree

console_output = open('console_output.txt', 'w')

username = input("Username: ")
while (username == ''):
    print('Username is blank, pleaes Enter Username: ')
    username = input("Username: ")
password = getpass.getpass(prompt=str("Password: "))
print('\n\n\n')
#************************Reading Excel File and Loading Data********
working_dir = os.getcwd() + str("\\")
print ('Your Working Directory: ' + working_dir, '\n\n')
xAPI_Report_read = openpyxl.load_workbook(working_dir + 'IP_address.xlsx')
xAPI_SheetName = xAPI_Report_read['IP_address']
macroNametoDelete = xAPI_SheetName.cell(2,2).value
interfaceNametoDelete = xAPI_SheetName.cell(2,3).value
print("You are deleting Macro: ", macroNametoDelete)
print("You are deleting UI Extensions: ", interfaceNametoDelete, '\n\n')
##**************************************************************************

##*******************xml data for xAPI_Macro******************************
xAPI_Macro = etree.Element("Command")
Macros = etree.SubElement(xAPI_Macro, "Macros")
Macro = etree.SubElement(Macros, "Macro")
RemoveMacro = etree.SubElement(Macro, "Remove")
Name = etree.SubElement(RemoveMacro, "Name").text ='' macroNametoDelete
xAPI_Macro_xml = etree.tostring(xAPI_Macro)
print(xAPI_Macro_xml)
##************************************************************************

##********************XML data for xAPI iNTERFACE***************************
xAPI_Interface = etree.Element("Command")
UserInterface = etree.SubElement(xAPI_Interface, "UserInterface")
Extensions = etree.SubElement(UserInterface, "Extensions")
Panel = etree.SubElement(Extensions, "Panel")
RemovePanelID = etree.SubElement(Panel, "Remove")
PanelID = etree.SubElement(RemovePanelID, "PanelID").text = interfaceNametoDelete
xAPI_Interface_xml = etree.tostring(xAPI_Interface)
#print(xAPI_Interface_xml)
#print(xAPI_Interface_xml, '\n\n')
##*************************************************************************
##################Connecting to device and Deleting macro and interface########
for i in range(2,xAPI_SheetName.max_row+1):
    print ('\n\n', xAPI_SheetName.cell(i,1).value)
    console_output.write('\n\n' + str(xAPI_SheetName.cell(i,1).value) +'\n')
    url = "http://"+xAPI_SheetName.cell(i,1).value + "/putxml"
    try:
        #########***************Deleting xAPI Macro*****************************
        payload = xAPI_Macro_xml
        headers = {'Content-Type': 'text/xml',}
        response = requests.request("POST", url, headers=headers, data = payload, auth=(username, password), verify=False)
        console_xml1 = (response.text.encode('utf8'), response.status_code)
        ###########************Deleting xAPI Interface***********************
        payload = xAPI_Interface_xml
        headers = {'Content-Type': 'application/xml',}
        response = requests.request("POST", url, headers=headers, data = payload, auth=(username, password), verify=False)
        console_xml2 = (response.text.encode('utf8'), response.status_code)
        
#        response1 = urlopen(requests("POST", url, headers=headers, data = payload, auth=(username, password), verify=False))
#        console_xml2 = (response1.read.decode('utf8'), response.status_code)
#        print(console_xml2)
#        json_object = json.loads(str(response.text.encode('utf8')))
#        print(json.dumps(json_object, sort_keys=True, indent=4))
        ###########Writing Output#######################################3
        console_date =  datetime.now()
        if "Unauthorized" in str(console_xml1) and response.status_code != 200 :
            print("Failed to login, please check user credentials!!\n")
            console_output.write("Failed to login, please check user credentials!!\n")
        else:
            print(console_xml1)
            print(console_xml2)
            print("Date and Time of Deletion: ", console_date)
            console_output.write(str(console_xml1) + '\n')
            console_output.write(str(console_xml2) + '\n')
            console_output.write("Date and Time of Deletion: " + str(console_date) + '\n')
                
    #########HTTP Error######################            
    except (requests.ConnectionError, requests.ConnectTimeout, requests.HTTPError) as e:
        print("Device is unreachable, please check if device is online!!\n\n")
        console_output.write("Device is unreachable, please check if device is online!!\n")
        console_output.write(str(e))
     
console_output.close()

check_exit = input("\n Type exit or close the session :")
while(check_exit != 'exit'):
    check_exit = input("Please type Exit: ")
print("Exiting in 5 sec") 
time.sleep(5)