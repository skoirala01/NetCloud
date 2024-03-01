import os
import shutil
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font


###################################################################################
log_dir = '<Directory>/'+time.strftime("%m-%d-%y")
readfile_frompath = 'Directory/DAV.xlsx'
################################################################
########Read or Load the Excel File#############################
filename = readfile_frompath
NBCUITT = load_workbook(filename)
sheet = NBCUITT.get_sheet_by_name('DAV')
#NBCUASA = Workbook()
#NBCUWLC = Workbook()
#NBCUBC = Workbook()

NBCUIT = load_workbook('<Directory>/Template_NBCU_IT.xlsx')
NBCUASA = load_workbook('<Directory>Template_NBCU_ASA.xlsx')
NBCUWLC = load_workbook('<Directory>Template_NBCU_WLC.xlsx')
NBCUBC = load_workbook('<Directory>Template_NBCU_Broadcast.xlsx')



################################################################
########Delete extra Coulumn####################################
sheet.delete_cols(18,19)
sheet.delete_cols(16,1)
sheet.delete_cols(14,1)
sheet.delete_cols(11,1)
sheet.delete_cols(9,1)
print (sheet.cell(4,10).value)
print (sheet.cell(4,11).value)
  
################################################################
########Create new Sheet that needed#############################
NBCUIT.create_sheet('IT',1)
#NBCUIT.create_sheet('Int`l',2)
NBCUIT.create_sheet('Reachable(IT)',2)
#NBCUIT.create_sheet('Reachable(Int`l)',4)
NBCUIT.create_sheet('UnReachable(IT)',3)
#NBCUIT.create_sheet('UnReachable(Int`l)',6)
NBCUIT.create_sheet('ConfigFailed(IT)',4)
#NBCUIT.create_sheet('ConfigFailed(Int`l)',8)
NBCUASA.create_sheet('ASA Inventory',1)
NBCUASA.create_sheet('ASA Reachable',2)
NBCUASA.create_sheet('ASA UnReachable',3)
NBCUASA.create_sheet('ASA ConfigFailed',4)
NBCUWLC.create_sheet('WLC Inventory',1)
NBCUWLC.create_sheet('WLC Reachable',2)
NBCUWLC.create_sheet('WLC UnReachable',3)
NBCUWLC.create_sheet('WLC ConfigFailed',4)
NBCUBC.create_sheet('BC Inventory',1)
NBCUBC.create_sheet('BC Reachable',2)
NBCUBC.create_sheet('BC UnReachable',3)
NBCUBC.create_sheet('BC ConfigFailed',4)

#NBCUIT.create_sheet('CoverPage',0)
#NBCUASA.create_sheet('CoverPage',0)
#NBCUWLC.create_sheet('CoverPage',0)
#NBCUBC.create_sheet('CoverPage',0)
################################################################
########Read the SheetName that are created######################
#it_coverpage = NBCUIT.get_sheet_by_name('CoverPage')
#asa_coverpage = NBCUASA.get_sheet_by_name('CoverPage')
#wlc_coverpage = NBCUWLC.get_sheet_by_name('CoverPage')
#bc_coverpage = NBCUBC.get_sheet_by_name('CoverPage')

us_coverpage = NBCUIT.get_sheet_by_name('CoverPage')
us_invent = NBCUIT.get_sheet_by_name('IT')
#intl_invent = NBCUIT.get_sheet_by_name('Int`l')
us_reach = NBCUIT.get_sheet_by_name('Reachable(IT)')
#intl_reach = NBCUIT.get_sheet_by_name('Reachable(Int`l)')
us_unreach = NBCUIT.get_sheet_by_name('UnReachable(IT)')
#intl_unreach = NBCUIT.get_sheet_by_name('UnReachable(Int`l)')
us_configfailed= NBCUIT.get_sheet_by_name('ConfigFailed(IT)')
#intl_configfailed = NBCUIT.get_sheet_by_name('ConfigFailed(Int`l)')

asa_coverpage = NBCUASA.get_sheet_by_name('CoverPage')
asa_invent = NBCUASA.get_sheet_by_name('ASA Inventory')
asa_reach = NBCUASA.get_sheet_by_name('ASA Reachable')
asa_unreach= NBCUASA.get_sheet_by_name('ASA UnReachable')
asa_configfailed= NBCUASA.get_sheet_by_name('ASA ConfigFailed')

wlc_coverpage = NBCUWLC.get_sheet_by_name('CoverPage')
wlc_invent = NBCUWLC.get_sheet_by_name('WLC Inventory')
wlc_reach = NBCUWLC.get_sheet_by_name('WLC Reachable')
wlc_unreach = NBCUWLC.get_sheet_by_name('WLC UnReachable')
wlc_configfailed= NBCUWLC.get_sheet_by_name('WLC ConfigFailed')

bc_coverpage = NBCUBC.get_sheet_by_name('CoverPage')
bc_invent = NBCUBC.get_sheet_by_name('BC Inventory')
bc_reach = NBCUBC.get_sheet_by_name('BC Reachable')
bc_unreach = NBCUBC.get_sheet_by_name('BC UnReachable')
bc_configfailed = NBCUBC.get_sheet_by_name('BC ConfigFailed')


#
################################################################
########Extract the US and International Inventory##############

row = sheet.max_row
col = sheet.max_column
a_row = 2
b_row = 2

intl = ['ams','auc','ban','bei','bel','ber','bru','dub','fra','ham',
        'hel','joh','lon','mad','mel','mex','mil','mos','mum',
        'mun','par','rom','sao','seo','sha','sin','sto','syd','tok','vie','zur',]
intl_len = len(intl)

BC = ['den-drydmz-fwx1','den-drydmz-fwx2','den-drymog-fw03','den-drymog-fw04','den-hubspk-fw01','den-hubspk-fw02',
'den-nocmux-ss08','den-nocmux-ss09','den-nocmux-ss11','eng-mog-fw01',
'SOLARWINDSAGENT','ucb-ucbdis-ts01','ucb-ucbdis-ts02',]

for j in range(1, col+1):
    us_invent.cell(1,j).value = sheet.cell(1,j).value    
#    intl_invent.cell(1,j).value = sheet.cell(1,j).value
    us_invent.cell(1,j).font = Font(sz = 12, b = True)
#    intl_invent.cell(1,j).font = Font(sz = 12, b = True)

for i in range (2,row+1):
   # print (sheet.cell(i,10).value)
#    check = 0
#    for it in range (0,intl_len):
#        if intl[it] not in str(sheet.cell(i,1).value):
#            check = 1
#        else:
#            check = 2
#            break
    for b in range (0,len(BC)):
        if BC[b] == str(sheet.cell(i,1).value):
            thisisBC = 'true'
        else:
            thisisBC = 'false'
    #if str(sheet.cell(i,10).value) == 'nbc2' and str(sheet.cell(i,11).value) != 'International':
    if (str(sheet.cell(i,11).value) == 'nbc2' and '-bc' not in str(sheet.cell(i,1).value) and '-BC' not in str(sheet.cell(i,1).value) and thisisBC == 'false' and '-hub' not in str(sheet.cell(i,1).value).lower() and '3rd Party' not in str(sheet.cell(i,6).value)):
        #if str((sheet.cell(i,12).value).lower() != 'rob'):# and check == 1:
        if 'rob' not in str(sheet.cell(i,12).value).lower():
           # print (sheet.cell(i,12).value).lower(), (type(sheet.cell(i,12).value))
            for j in range (1,col+1):
                us_invent.cell(a_row,j).value = sheet.cell(i,j).value
                #print sheet.cell(i,j).value
            a_row = a_row + 1
        
#    #elif str(sheet.cell(i,11).value) == 'International':
#    elif str(sheet.cell(i,11).value) == 'nbc2' and check == 2 and '-bc' not in str(sheet.cell(i,1).value) and '-BC' not in str(sheet.cell(i,1).value) and '3rd Party' not in str(sheet.cell(i,6).value):
#        for j in range (1,col+1):
#            intl_invent.cell(b_row,j).value = sheet.cell(i,j).value
#        b_row = b_row + 1

################################################################
########Extract US Reachable and Unreachable####################
row = us_invent.max_row
col = us_invent.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    us_reach.cell(1,j).value = us_invent.cell(1,j).value
    us_unreach.cell(1,j).value = us_invent.cell(1,j).value
    us_reach.cell(1,j).font = Font(sz = 12, b = True)
    us_unreach.cell(1,j).font = Font(sz = 12, b = True)

for i in range (2,row+1):
    if str(us_invent.cell(i,3).value) == 'ACTIVE':
        for j in range (1,col+1):
            us_reach.cell(a_row,j).value = us_invent.cell(i,j).value
            #print us_invent.cell(i,j).value
        a_row = a_row + 1
        
    elif str(us_invent.cell(i,3).value) == 'DEVICE NOT REACHABLE':
        for j in range (1,col+1):
            us_unreach.cell(b_row,j).value = us_invent.cell(i,j).value
            #print us_invent.cell(i,j).value
        b_row = b_row + 1

################################################################
########Extract Internaltional Rechable and Unreachable##########
#row = intl_invent.max_row
#col = intl_invent.max_column
#a_row = 2
#b_row = 2
#for j in range(1, col+1):
#    intl_reach.cell(1,j).value = intl_invent.cell(1,j).value
#    intl_unreach.cell(1,j).value = intl_invent.cell(1,j).value
#    intl_reach.cell(1,j).font = Font(sz = 12, b = True)
#    intl_unreach.cell(1,j).font = Font(sz = 12, b = True)
#
#for i in range (2,row+1):
#    if str(intl_invent.cell(i,3).value) == 'ACTIVE':
#        for j in range (1,col+1):
#            intl_reach.cell(a_row,j).value = intl_invent.cell(i,j).value
#        a_row = a_row + 1
#        
#    elif str(intl_invent.cell(i,3).value) == 'DEVICE NOT REACHABLE':
#        for j in range (1,col+1):
#            intl_unreach.cell(b_row,j).value = intl_invent.cell(i,j).value
#        b_row = b_row + 1

################################################################
########Extract Config failed for US and International##########
row = us_reach.max_row
col = us_reach.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    us_configfailed.cell(1,j).value = us_reach.cell(1,j).value
#    intl_configfailed.cell(1,j).value = intl_reach.cell(1,j).value
    us_configfailed.cell(1,j).font = Font(sz = 12, b = True)
#    intl_configfailed.cell(1,j).font = Font(sz = 12, b = True)
    

for i in range (2,row+1):
    if str(us_reach.cell(i,14).value) != 'PASS':
        for j in range (1,col+1):
            us_configfailed.cell(a_row,j).value = us_reach.cell(i,j).value
        a_row = a_row + 1

#row = intl_reach.max_row
#for i in range (2,row+1):
#    if str(intl_reach.cell(i,14).value) != 'PASS':
#        for j in range (1,col+1):
#            intl_configfailed.cell(b_row,j).value = intl_reach.cell(i,j).value
#        b_row = b_row + 1
        
###################################################################
#######################ASA and WLC Inventory###################################
row = sheet.max_row
col = sheet.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    asa_invent.cell(1,j).value = sheet.cell(1,j).value
    asa_invent.cell(1,j).font = Font(sz = 12, b = True)
    wlc_invent.cell(1,j).value = sheet.cell(1,j).value
    wlc_invent.cell(1,j).font = Font(sz = 12, b = True)

for i in range (2,row+1):
    if 'ASA' in str(sheet.cell(i,7).value):
        for j in range (1,col+1):
            asa_invent.cell(a_row,j).value = sheet.cell(i,j).value
        a_row = a_row + 1
    elif 'Wireless' in str(sheet.cell(i,7).value):
        for j in range (1,col+1):
            wlc_invent.cell(b_row,j).value = sheet.cell(i,j).value
        b_row = b_row + 1

################################################################
########ASA reachable unreachable ##############################
row = asa_invent.max_row
col = asa_invent.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    asa_reach.cell(1,j).value = asa_invent.cell(1,j).value
    asa_unreach.cell(1,j).value = asa_invent.cell(1,j).value
    asa_reach.cell(1,j).font = Font(sz = 12, b = True)
    asa_unreach.cell(1,j).font = Font(sz = 12, b = True)

for i in range (2,row+1):
    if str(asa_invent.cell(i,3).value) == 'ACTIVE':
        for j in range (1,col+1):
            asa_reach.cell(a_row,j).value = asa_invent.cell(i,j).value
        a_row = a_row + 1
        
    elif str(asa_invent.cell(i,3).value) == 'DEVICE NOT REACHABLE':
        for j in range (1,col+1):
            asa_unreach.cell(b_row,j).value = asa_invent.cell(i,j).value
        b_row = b_row + 1
################################################################
########WLC reachable unreachable ##############################
row = wlc_invent.max_row
col = wlc_invent.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    wlc_reach.cell(1,j).value = wlc_invent.cell(1,j).value
    wlc_unreach.cell(1,j).value = wlc_invent.cell(1,j).value
    wlc_reach.cell(1,j).font = Font(sz = 12, b = True)
    wlc_unreach.cell(1,j).font = Font(sz = 12, b = True)

for i in range (2,row+1):
    if str(wlc_invent.cell(i,3).value) == 'ACTIVE':
        for j in range (1,col+1):
            wlc_reach.cell(a_row,j).value = wlc_invent.cell(i,j).value
        a_row = a_row + 1
        
    elif str(asa_invent.cell(i,3).value) == 'DEVICE NOT REACHABLE':
        for j in range (1,col+1):
            wlc_unreach.cell(b_row,j).value = wlc_invent.cell(i,j).value
        b_row = b_row + 1
################################################################
########Extract Config failed for ASA and WLC##########
row = asa_reach.max_row
col = asa_reach.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    asa_configfailed.cell(1,j).value = asa_reach.cell(1,j).value
    wlc_configfailed.cell(1,j).value = wlc_reach.cell(1,j).value
    asa_configfailed.cell(1,j).font = Font(sz = 12, b = True)
    wlc_configfailed.cell(1,j).font = Font(sz = 12, b = True)
    

for i in range (2,row+1):
    if str(asa_reach.cell(i,14).value) != 'PASS':
        for j in range (1,col+1):
            asa_configfailed.cell(a_row,j).value = asa_reach.cell(i,j).value
        a_row = a_row + 1

row = wlc_reach.max_row
col = asa_reach.max_column
for i in range (2,row+1):
    if str(wlc_reach.cell(i,14).value) != 'PASS':
        for j in range (1,col+1):
            wlc_configfailed.cell(b_row,j).value = wlc_reach.cell(i,j).value
        b_row = b_row + 1
        
###################################################################
#######################Extract BC or bc Inventory###################################


row = sheet.max_row
col = sheet.max_column
a_row = 2
for j in range(1, col+1):
    bc_invent.cell(1,j).value = sheet.cell(1,j).value

for i in range (2,row+1):
    checkbc = 0
    for b in range (0,len(BC)):
        if BC[b] == str(sheet.cell(i,1).value) and checkbc == 0 and '3rd Party' not in str(sheet.cell(i,6).value):
            checkbc = 1
            break
        else:
            checkbc = 0

    if ('-bc' in str(sheet.cell(i,1).value) or '-BC' in str(sheet.cell(i,1).value) or '-HUB' in str(sheet.cell(i,1).value) or '-hub' in str(sheet.cell(i,1).value)) and '3rd Party' not in str(sheet.cell(i,6).value) or 'rob' in str(sheet.cell(i,12).value).lower():
        for j in range (1,col+1):
            bc_invent.cell(a_row,j).value = sheet.cell(i,j).value
        a_row = a_row + 1
    elif checkbc == 1:
        for j in range (1,col+1):
            bc_invent.cell(a_row,j).value = sheet.cell(i,j).value
            checkbc = 0
        a_row = a_row + 1
                

################################################################
########Broadcast reachable unreachable ##############################
row = bc_invent.max_row
col = bc_invent.max_column
a_row = 2
b_row = 2
for j in range(1, col+1):
    bc_reach.cell(1,j).value = asa_invent.cell(1,j).value
    bc_unreach.cell(1,j).value = asa_invent.cell(1,j).value
    bc_reach.cell(1,j).font = Font(sz = 12, b = True)
    bc_unreach.cell(1,j).font = Font(sz = 12, b = True)

for i in range (2, row+1):
    if str(bc_invent.cell(i,3).value) == 'ACTIVE':
        for j in range (1,col+1):
            bc_reach.cell(a_row,j).value = bc_invent.cell(i,j).value
        a_row = a_row + 1
    elif str(bc_invent.cell(i,3).value) == 'DEVICE NOT REACHABLE':
        for j in range (1,col+1):
            bc_unreach.cell(b_row,j).value = bc_invent.cell(i,j).value
        b_row = b_row + 1

################################################################
########Extract Config failed for BC##########
row = bc_reach.max_row
col = bc_reach.max_column
a_row = 2
for j in range(1, col+1):
    bc_configfailed.cell(1,j).value = bc_reach.cell(1,j).value
    bc_configfailed.cell(1,j).font = Font(sz = 12, b = True)
   

for i in range (2,row+1):
    if str(bc_reach.cell(i,14).value) != 'PASS':
        for j in range (1,col+1):
            bc_configfailed.cell(a_row,j).value = bc_reach.cell(i,j).value
        a_row = a_row + 1

#############################################################################
################################################################################

################################################################
########Counting total Device from Each Sheet###################
row = sheet.max_row-1
row_us = us_invent.max_row-1
#row_int = intl_invent.max_row-1
row_us_reach = us_reach.max_row-1
#row_int_reach = intl_reach.max_row-1
row_us_unreach = us_unreach.max_row-1
#row_int_unreach = intl_unreach.max_row-1
row_us_configfailed = us_configfailed.max_row-1
#row_int_configfailed = intl_configfailed.max_row-1
row_asa  = asa_invent.max_row-1
row_asa  = asa_invent.max_row-1
row_asa_reach = asa_reach.max_row-1
row_asa_configfailed = asa_configfailed.max_row-1
row_asa_unreach = asa_unreach.max_row-1
row_wlc = wlc_invent.max_row-1
row_wlc_reach = wlc_reach.max_row-1
row_wlc_unreach = wlc_unreach.max_row-1
row_wlc_configfailed = wlc_configfailed.max_row-1
row_bc = bc_invent.max_row-1
row_bc_reach = bc_reach.max_row-1
row_bc_unreach = bc_unreach.max_row-1
row_bc_configfailed = bc_configfailed.max_row-1


################copy values in cover page############
#it_coverpage.cell(10,3).value = row_us
#it_coverpage.cell(11,3).value = row_us_reach
#it_coverpage.cell(12,3).value = row_us_configfailed
#it_coverpage.cell(13,3).value = row_us_unreach
#
#asa_coverpage.cell(8,3).value = row_asa
#asa_coverpage.cell(9,3).value = row_asa_reach
#asa_coverpage.cell(10,3).value = row_asa_configfailed
#asa_coverpage.cell(11,3).value = row_asa_unreach
#
#wlc_coverpage.cell(8,3).value = row_wlc
#wlc_coverpage.cell(9,3).value = row_wlc_reach
#wlc_coverpage.cell(10,3).value = row_wlc_configfailed
#wlc_coverpage.cell(11,3).value = row_wlc_unreach
#
#bc_coverpage.cell(8,3).value = row_bc
#bc_coverpage.cell(9,3).value = row_bc_reach
#bc_coverpage.cell(10,3).value = row_bc_configfailed
#bc_coverpage.cell(11,3).value = row_bc_unreach

################################################################
########Renaming the sheet######################################
sheet.title = 'Inventory -' + str(row)
us_invent.title = us_invent.title + '-Invent-'+str(row_us)
#intl_invent.title = intl_invent.title + '-Invent-'+str(row_int)
us_reach.title = us_reach.title + '-'+str(row_us_reach)
#intl_reach.title = intl_reach.title + '-'+str(row_int_reach)
us_unreach.title = us_unreach.title + '-'+str(row_us_unreach)
#intl_unreach.title = intl_unreach.title + '-'+str(row_int_unreach)
us_configfailed.title = us_configfailed.title + '-' + str(row_us_configfailed)
#intl_configfailed.title = intl_configfailed.title + '-' + str(row_int_configfailed)
asa_invent.title = asa_invent.title + '-' + str(row_asa)
asa_reach.title = asa_reach.title + '-' + str(row_asa_reach)
asa_unreach.title = asa_unreach.title + '-' + str (row_asa_unreach)
asa_configfailed.title = asa_configfailed.title + '-' + str (row_asa_configfailed)
wlc_invent.title = wlc_invent.title + '-' + str(row_wlc)
wlc_reach.title = wlc_reach.title + '-' + str(row_wlc_reach)
wlc_unreach.title = wlc_unreach.title + '-' + str (row_wlc_unreach)
wlc_configfailed.title = wlc_configfailed.title + '-' + str (row_wlc_configfailed)
bc_invent.title = bc_invent.title + '-' + str(row_bc)
bc_reach.title = bc_reach.title + '-' + str(row_bc_reach)
bc_unreach.title = bc_unreach.title + '-' + str (row_bc_unreach)
bc_configfailed.title = bc_configfailed.title + '-' + str (row_bc_configfailed)

####################################CoverPage############################################
#us_coverpage.insert_cols(3)
#bc_coverpage.insert_cols(3)
#asa_coverpage.insert_cols(2)
#wlc_coverpage.insert_cols(2)

us_coverpage.cell(9,3).value = time.strftime("%m-%d-%y")
bc_coverpage.cell(7,3).value = time.strftime("%m-%d-%y")
asa_coverpage.cell(2,2).value = time.strftime("%m-%d-%y")
wlc_coverpage.cell(2,2).value = time.strftime("%m-%d-%y")

us_coverpage.cell(10,3).value = row_us
us_coverpage.cell(11,3).value = row_us
us_coverpage.cell(13,3).value = row_us_reach
us_coverpage.cell(14,3).value = row_us_reach
us_coverpage.cell(16,3).value = row_us_configfailed
us_coverpage.cell(17,3).value = row_us_configfailed
us_coverpage.cell(19,3).value = row_us_unreach
us_coverpage.cell(20,3).value = row_us_unreach

bc_coverpage.cell(8,3).value = row_bc
bc_coverpage.cell(9,3).value = row_bc_reach
bc_coverpage.cell(10,3).value = row_bc_configfailed
bc_coverpage.cell(11,3).value = row_bc_unreach


asa_coverpage.cell(3,2).value = row_asa
asa_coverpage.cell(4,2).value = row_asa_reach
asa_coverpage.cell(5,2).value = row_asa_configfailed
asa_coverpage.cell(6,2).value = row_asa_unreach

wlc_coverpage.cell(3,2).value = row_wlc
wlc_coverpage.cell(4,2).value = row_wlc_reach
wlc_coverpage.cell(5,2).value = row_wlc_configfailed
wlc_coverpage.cell(6,2).value = row_wlc_unreach

#####################################################################################
#######################################################################################
file_asa = 'NBCU_ASA_reachable_unrechable_count_'+time.strftime("%m-%d-%y" + '.xlsx')
file_wlc = 'NBCU_WLC_reachable_unreachable_count_'+time.strftime("%m-%d-%y" + '.xlsx')
file_bc = 'Broadcast_reachable_unreachable_count_'+time.strftime("%m-%d-%y" + '.xlsx')

NBCUIT.save(filename)
NBCUASA.save(file_asa)
NBCUWLC.save(file_wlc)
NBCUBC.save(file_bc)




try:
    os.makedirs(log_dir)
except:
    print("Directory already exists, removing and creating it")
    shutil.rmtree(log_dir)
    os.makedirs(log_dir)

newfilename = 'NBCU_IT_reachable_unreachable_count_' + time.strftime("%m-%d-%y") + '.xlsx'
os.rename(filename, newfilename)
shutil.copy2(newfilename, log_dir)
shutil.copy2(file_asa, log_dir)
shutil.copy2(file_wlc, log_dir)
shutil.copy2(file_bc, log_dir)
os.remove(newfilename)
os.remove(file_asa)
os.remove(file_wlc)
os.remove(file_bc)

