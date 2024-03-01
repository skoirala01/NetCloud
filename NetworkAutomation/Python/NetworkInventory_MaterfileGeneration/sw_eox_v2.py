import os
import re
import shutil
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
import win32com.client

def nbcu_sw_eox(path_to_read,path_to_save):
    

 
 ####creating new file for LDOS
    SW_EOX = Workbook()
    SW_EOX.create_sheet("NBCU_SW_Planning",0)
    sw_eox = SW_EOX.get_sheet_by_name('NBCU_SW_Planning')
   
    filename = path_to_read#'C:/Users/skoirala/Downloads/TRACK_DETAILS_20181115_1330.xlsx'
    sw_track_files = load_workbook(filename)    
    sheetsNamess = (sw_track_files.worksheets)
    sheetCount =  len(sheetsNamess)
    
    sheetsNames = (sw_track_files.worksheets[3])
    for j in range(1, 4):
        sw_eox.cell(1,j).value = sheetsNames.cell(10,j).value
        sw_eox.cell(1,j).font = Font(sz = 12, b = True)
    for j in range(9, 13):
       sw_eox.cell(1,j).value = sheetsNames.cell(10,j+6).value
       sw_eox.cell(1,j).font = Font(sz = 12, b = True)
        
    sw_eox.cell(1,4).value = 'Platform SW Recommendation(N, Absolute Conformance)'
    sw_eox.cell(1,4).font = Font(sz = 12, b = True)
    sw_eox.cell(1,5).value = 'Date Of SW Recommendation'
    sw_eox.cell(1,5).font = Font(sz = 12, b = True)
    sw_eox.cell(1,6).value = 'Platform SW Recommendation(N-1, Flexible Confromance)'
    sw_eox.cell(1,6).font = Font(sz = 12, b = True)
    sw_eox.cell(1,7).value = 'Current  SW version in Production'
    sw_eox.cell(1,7).font = Font(sz = 12, b = True)
    sw_eox.cell(1,8).value = 'SW Conformance'
    sw_eox.cell(1,8).font = Font(sz = 12, b = True)  
    sw_eox.cell(1,13).value = 'Hardware EoSWM Notes'
    sw_eox.cell(1,13).font = Font(sz = 12, b = True)      
      
    
    

    a_row = 2
    for count in range (3,sheetCount):
        #print 10000
        sheetsNames = ((sw_track_files.worksheets[count]))
        row = sheetsNames.max_row
#        print row
#        sp = str(sheetsNames.title)
#        print type(sp)
#       
#        print sp
#        print count
#        
        if ('IOS_XR_Cisco_ASR_9000_' not in str(sheetsNames.title) and 'IOS_XR_Cisco_ASR_990x' not in str(sheetsNames.title) and 'IOS_XR_NCS540_PE' not in str(sheetsNames.title)):
            
            for i in range (11,row+1):
#                print i
                for j in range (1,4):##get first 3 colns from SW TRACK
                    sw_eox.cell(a_row,j).value = sheetsNames.cell(i,j+1).value    
#                    print sw_eox.cell(a_row,j).value 
#                    print sheetsNames.cell(i,j+1).value
                sw_eox.cell(a_row,4).value = sheetsNames.cell(8,4).value#get the current Reccomended sw version(n)
#                print(sheetsNames.cell(8,4).value)
                sw_eox.cell(a_row,5).value = sheetsNames.cell(9,4).value#get the date of recommendation
                sw_eox.cell(a_row,6).value = sheetsNames.cell(8,6).value#get previous recommended code version(n-1)
                sw_eox.cell(a_row,7).value = sheetsNames.cell(i,6).value##get the production version
#                print(sheetsNames.cell(8,6).value)
                N_1 = str((sheetsNames.cell(8,6).value).strip())#get previous recommended code version(n-1)
#                print(sheetsNames.cell(i,6).value)
                print((sheetsNames.cell(i,6).value))
                N_prod = str((sheetsNames.cell(i,6).value).strip())#get the production version
                N = str((sheetsNames.cell(8,4).value).strip())#get the current Reccomended sw version(n)
                
                if sheetsNames.cell(i,1).value == 'YES':
                    sw_eox.cell(a_row,8).value = 'Absolute Conformance'#sheetsNames.cell(i,1).value##get confirmance information
                
                elif (str(sheetsNames.cell(i,1).value)).lower() == "no" and N_1 == N_prod: #checking flexible conformance
                    sw_eox.cell(a_row,8).value = 'Flexible Conformance'

                ######IOSXE Checking###############################               
                elif str(N_prod) > str(N):#checking higher than reccommended code version
                   # print sw_eox.cell(a_row,8).value
                    sw_eox.cell(a_row,8).value = 'NO'
                    N_prod_countdot = N_prod.count('.')
                    N_countdot = N.count('.')
                    
                    if N_prod_countdot >= 2:#########checking IOS XE
                        #N_1_split = N_1.split('.')
                        N_prod_split = N_prod.split('.')
                        N_split = N.split('.')
                        if N_prod_countdot >= 2 and N_countdot >= 2:
                            if N_prod_split[0] == N_split[0] and N_prod_split[1] == N_split[1]:
                                if N_prod_split[2] > N_split[2]:
                                    sw_eox.cell(a_row,8).value = 'Higher than N version in the same release train'
#                    if N_prod_countdot >= 3:#########checking IOS XE
#                    #N_1_split = N_1.split('.')
#                        N_prod_split = N_prod.split('.')
#                        N_split = N.split('.')
#                        if N_prod_countdot >= 2 and N_countdot >= 2:
#                            if N_prod_split[0] == N_split[0] and N_prod_split[1] == N_split[1]:
#                                if N_prod_split[2] > N_split[2]:
#                                    sw_eox.cell(a_row,8).value = 'Higher than N version in the same release train'
                                
                    elif N_prod_countdot < 2 and N_countdot < 2 and N_prod != 'Not Found':########checking IOS softwares
                        #############Current Recom########
                        N_split = N.split(')')
                        N_1st = N_split[0]
                        N_1st = N_1st.split('(')[0]
                        if N_split[1] != '':
                            N_2nd = N_split[1]
                            check_last_digit = []
                            for k in range(0, len(N_2nd)):
                                if N_2nd[k] in '0123456789':
                                    check_last_digit.append(N_2nd[k])
                            N_lastdigit = ""
                            N_lastdigit = N_lastdigit.join(check_last_digit)
                            if N_lastdigit != '':
                                N_lastchar = N_2nd.split(N_lastdigit)[0]
                                N_lastdigit = int(N_lastdigit)
                        ##############################N-1 secition####
                        N_prod_split = N_prod.split(')')
                        N_prod_1st = N_prod_split[0]
                        N_prod_1st = N_prod_1st.split('(')[0]
                        #print N_prod_1st,3
                        if N_prod_1st != '':
                            N_prod_2nd = N_prod_split[1]
                            check_last_digit = []
                            for k in range(0, len(N_prod_2nd)):
                                if N_prod_2nd[k] in '0123456789':
                                    check_last_digit.append(N_prod_2nd[k])
                            N_prod_lastdigit = ""
                            N_prod_lastdigit = N_prod_lastdigit.join(check_last_digit)
                            if N_prod_lastdigit != '':
                                N_prod_lastchar = N_prod_2nd.split(N_prod_lastdigit)[0]
                                N_prod_lastdigit = int(N_prod_lastdigit)
                        
                        if N_prod_1st == N_1st and N_prod_lastchar == N_lastchar and N_prod_lastdigit > N_lastdigit:
                            sw_eox.cell(a_row,8).value = 'Higher than N version in the same release train'
                    
                else:
                    sw_eox.cell(a_row,8).value = 'NO'
                a_row = a_row + 1
                        
                
                
        elif  'IOS_XR_Cisco_ASR_9000' in sheetsNames.title or 'IOS_XR_Cisco_ASR_990x' in str(sheetsNames.title) or 'IOS_XR_NCS540_PE' in str(sheetsNames.title):         
            for i in range (14,row+1):
                
                for j in range (1,4):
                    sw_eox.cell(a_row,j).value = sheetsNames.cell(i,j+1).value        
                sw_eox.cell(a_row,4).value = sheetsNames.cell(9,6).value#recommneded version (n)
                sw_eox.cell(a_row,5).value = sheetsNames.cell(9,5).value#Recommended date
                sw_eox.cell(a_row,6).value = sheetsNames.cell(10,6).value#previous recommendation(n-1)
                sw_eox.cell(a_row,7).value = sheetsNames.cell(i,6).value#production version
                #print(sheetsNames.cell(10,6).value)
                N_1 = str((sheetsNames.cell(10,6).value).strip())#recommneded version (n-1)
                N_prod = str((sheetsNames.cell(i,6).value).strip())#production version
                N = str((sheetsNames.cell(9,6).value).strip())#get the current Reccomended sw version(n)
   
                if sheetsNames.cell(i,1).value == 'YES' or N_prod == N:
                    sw_eox.cell(a_row,8).value = 'Absolute Conformance'#sheetsNames.cell(i,1).value##get confirmance information
                elif sheetsNames.cell(i,1).value == 'NO' and N_prod == N_1:
                    sw_eox.cell(a_row,8).value = 'Flexible Conformance'

#                elif str(N_prod) > str(N):#checking higher than reccommended code version
#                    sw_eox.cell(a_row,8).value = "Higher than N version in the same release train"                    

                elif sheetsNames.cell(i,1).value == 'NO' and (N_prod != N_1 or N_prod != N):
                    sw_eox.cell(a_row,8).value = 'NO'

                for j in range(9,13):
                    sw_eox.cell(a_row,j).value = sheetsNames.cell(i,j+8).value#copy of the sw planning                  
                sw_eox.cell(a_row,13).value = sheetsNames.cell(i,22).value#copy last tab
                a_row = a_row + 1

#######################################################
                
###################Deleting Duplicates##################
#    
    duplicate_count = 0
    row = sw_eox.max_row
    print (row)
    for i in range(2,row+1):
        if (sw_eox.cell(i,1).value) != None:
            for j in range(i+1,row+1):
                if str(sw_eox.cell(i,1).value) == str(sw_eox.cell(j,1).value):             
                    duplicate_count = duplicate_count + 1
                    sw_eox.delete_rows(j,1)
                    print ("\n\n This is Duplicate and Deleting it "+str((sw_eox.cell(j,1).value)))


####################Filter SUP6E#######################################
    SUP6E = ['chi-dlt1fl-as01','chi-dlt1fl-as02','eng-cnbcnm-ss01','eng-cnbcnm-ss02','nyc-30r05e-as01','nyc-30r10e-as01',
             'nyc-30r10e-as02','nyc-30r10w-as01','nyc-30r15e-av01','nyc-30r16e-av01','nyc-30r18e-as01','nyc-30r21e-as01',
             'rom-viapo-cs01','ucs-522501-as01','wdc-wrc111dc-as01','wdc-wrc3fl-as01','wdc-wrcpoo-as01',]
    #print len(SUP6E)
    for i in range(0, len(SUP6E)):
        for j in range(2,(row-duplicate_count+1)):
#            print (sw_eox.cell(j,3).value)
#            print j
            if SUP6E[i] in sw_eox.cell(j,1).value:#  and type(sw_eox.cell(j,3).value) != None:
                if '-E' in str(sw_eox.cell(j,3).value):
                    sw_eox.cell(j,8).value = 'NO /SUP6E'
                    print (sw_eox.cell(a_row,8).value)
##########################################################################
####################Filter Nexus7K 4Gigmem with SUP1 #######################################
    N7KSUP1 = ['cha-affexc-lr01','chi-wmaqit-cr01','chi-wmaqit-cr02','lon-centra-cr01','lon-centra-cr02','nyc-1221aa-cr01',
    'nyc-1221aa-cr02','nyc-32aint-cr01','nyc-32aint-cr02','nyc-32aofa-cr01','nyc-32aofa-cr02','nyc-5timsq-lr01',
    'nyc-5timsq-lr02','nyc-affexc-lr01','sta-pocadm-lr01','sta-pocadm-lr02','sta-pocavd-lr01','sta-pocavd-lr02',
    'sta-poccor-cr01','sta-poccor-cr02','sta-pocprd-lr01','sta-pocprd-lr02','ucs-1280dc-cr01','ucs-1360dc-cr01',
    'ucs-1360nx-sd01','ucs-1360nx-sd02','ucs-1360nx-ss01','ucs-1360nx-ss02']

    #print len(SUP6E)
    for i in range(0, len(N7KSUP1)):
        for j in range(2,(row-duplicate_count+1)):
#            print (sw_eox.cell(j,3).value)
#            print j
            if N7KSUP1[i] in sw_eox.cell(j,1).value:#  and type(sw_eox.cell(j,3).value) != None:
                if 'N7K' in str(sw_eox.cell(j,3).value):
                    sw_eox.cell(j,8).value = 'NO /Nexus7K 4Gigmem with SUP1'
                    print (sw_eox.cell(a_row,8).value)
                    
##########################################################################

    sw_eox.title = sw_eox.title + "_ " + str(row-(duplicate_count/2))
    newfilename = path_to_save+ 'NBCU_SW_Planning_' + time.strftime("%m-%d-%y") + '.xlsx'
    SW_EOX.save(newfilename)
    print ("\n\n\n SIAR Report 'NBCU_SW_Planning' is Saved, Total = " + str(row-(duplicate_count/2)) + "\n\n\n")
   