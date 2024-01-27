#from site-packages.toolkit.interface import interface
import os
import shutil
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
import win32com.client
import sw_eox_v2 as sweox
import hw_eox as hweox



#print ("Please provide date of last report in the format of MM-DD-YY")
date_lastreport  = input("Please provide date of last report in the format of MM-DD-YY:: ")
#print ((date_lastreport))

#read the proper paths to each file
NBCU_Inventory_Path_to_Date = 'C:/Users/skoirala/Documents/Works/1. NBCU/3. CSPC INVENTORY DELIVERABLES/' + time.strftime("%m-%d-%y")
path_to_read_siar = 'C:/Users/skoirala/Downloads/TRACK_DETAILS.xlsx'
path_to_read_eox = 'C:/Users/skoirala/Downloads/HW_EOX_DETAILS.xlsx'
path_to_save = 'C:/Users/skoirala/Documents/Works/1. NBCU/EOX and SIAR/'##please save 'NBCU_Steve_Update.xlsx'
SearchReport = 'C:/Users/skoirala/Downloads/77770_search_results.xlsx'


#call the sw_eox function
sweox.nbcu_sw_eox(path_to_read_siar,path_to_save)
#cal the hw_eox function
hweox.nbcu_hw_eox(path_to_read_eox,path_to_save)

#

############################################################################################3
#####Load the MasterReport Template##########
print("\n\n Loading the MasterReport Template \n\n")
MasterReportTemplate = 'NBCU_SW_HW_LDOS_Template.xlsx'
MasterReportSave = path_to_save + MasterReportTemplate
MasterReport = load_workbook(MasterReportSave)
mastersheets = MasterReport.worksheets
mastersheet = mastersheets[1]
master_row_start = 3

####Load the SIAR#######################################################################
print("\n\n Loading SIAR report recently saved in your folder\n\n")
SIARFileName = 'NBCU_SW_Planning_' + time.strftime("%m-%d-%y") + '.xlsx'
SIARFilePath = path_to_save + SIARFileName
SIARReport = load_workbook(SIARFilePath)
siarsheets = SIARReport.worksheets
siarsheet = siarsheets[0]
row_siar = siarsheet.max_row

####copy SIAR####
print("\n\n Copying SIAR report to Mater Template \n\n")
for i in range(2,row_siar+1):
    for j in range(2,5):
        mastersheet.cell(master_row_start,j).value = siarsheet.cell(i,j-1).value
    for j in range(6,11):
        mastersheet.cell(master_row_start,j).value = siarsheet.cell(i,j-2).value
    for j in range(12,16):
        mastersheet.cell(master_row_start,j).value = siarsheet.cell(i,j-3).value
    mastersheet.cell(master_row_start,20).value = siarsheet.cell(i,13).value
    
    master_row_start = master_row_start+1
        
SIARReport = ''
siarsheets = ''
    
####Load the EOX###############################################################
print("\n\n Loading EOX report  \n\n")
EOXFileName = 'NBCU_HW_Planning_LDOS_' + time.strftime("%m-%d-%y") + '.xlsx'
EOXFilePath = path_to_save + EOXFileName
EOXReport = load_workbook(EOXFilePath)
eoxsheets = EOXReport.worksheets
eoxsheet = eoxsheets[0]
row_eox = eoxsheet.max_row

##Copy EOX
print("\n\n Copying EOX report to the MasterReport Template \n\n")
for i in range (3,master_row_start+1):#loop for sw template
    for j in range(2,row_eox+1):#loop for hw template
        if mastersheet.cell(i,2).value == eoxsheet.cell(j,3).value:
            mastersheet.cell(i,4).value = eoxsheet.cell(j,1).value#get the pid
            mastersheet.cell(i,5).value = eoxsheet.cell(j,6).value#get the SN
            mastersheet.cell(i,9).value = eoxsheet.cell(j,5).value#get production version
            for k in range(1,5):
                mastersheet.cell(i,k+15).value = eoxsheet.cell(j,k+6).value
          
            value1 = eoxsheet.cell(j,3).value
            value2 = eoxsheet.cell(j+1,3).value
            eoxsheet.cell(j,3).value = 'copied'####no more hostname in eox to current row
            while(value1 == value2):
                i = i + 1
                j = j + 1
                master_row_start = master_row_start +1
                mastersheet.insert_rows(i)
                
                ##get individual first
                mastersheet.cell(i,4).value = eoxsheet.cell(j,1).value#get the pid
                mastersheet.cell(i,5).value = eoxsheet.cell(j,6).value#get SN
                mastersheet.cell(i,9).value = eoxsheet.cell(j,5).value#get production version
                
                for k in range(1,5):##copy of the multichasis from eox to master
                    mastersheet.cell(i,k+15).value = eoxsheet.cell(j,k+6).value
                for k in range(2,4):#copy of hostname and chasis name in new row
                    mastersheet.cell(i,k).value = mastersheet.cell(i-1,k).value
                for k in range(6,9):
                    mastersheet.cell(i,k).value = mastersheet.cell(i-1,k).value
                    

                
                value1 = eoxsheet.cell(j,3).value
                value2 = eoxsheet.cell(j+1,3).value
                eoxsheet.cell(j,3).value = 'copied'
EOXReport = ''
eoxsheets = ''                         
                
################Tag to the BC and International Devices################################
print("\n\n Tagging BC  and International Devices \n\n")
row_sw = mastersheet.max_row
for i in range (3,row_sw+1):
    hostname = str(mastersheet.cell(i,2).value)
    
    if '-bc' in hostname:
        mastersheet.cell(i,22).value = 'yes'
    else:
        mastersheet.cell(i,22).value = 'no'
 
      
intl = ['ams','auc','ban','bei','bel','ber','bru','dub','fra','ham',
        'hel','joh','lon','mad','mel','mex','mil','mos','mum',
        'mun','par','rom','sao','seo','sha','sin','sto','syd','tok','vie','zur',]

for i in range (3,row_sw+1):
    hostname = str(mastersheet.cell(i,2).value)
    for j in range(1,len(intl)):
        if (intl[j] in hostname):
            mastersheet.cell(i,23).value = 'yes'
            break
        else:
            mastersheet.cell(i,23).value = 'no'

###########################Update the Serial Number from Search Report in NP#############################
row_sw = mastersheet.max_row
print ("\n Loading and Updating SN from the Search Report...\n\n"       )
SerialNumber_SearchReport = load_workbook(SearchReport)
serialnumber = SerialNumber_SearchReport.get_sheet_by_name('77770_search_results')
row_sn = serialnumber.max_row
for i in range(3,row_sw+1):
    for j in range(1,row_sn+1):
        if mastersheet.cell(i,2).value == serialnumber.cell(j,1).value:
            if serialnumber.cell(j,8).value == None and mastersheet.cell(i,5).value == None:
                mastersheet.cell(i,4).value = 'None'
            elif serialnumber.cell(j,8).value != None and mastersheet.cell(i,5).value == None:
                mastersheet.cell(i,5).value = serialnumber.cell(j,8).value
            

SerialNumber_SearchReport = ''
serialnumber = ''

############################################################################################
####################Conditional Flexible Conformance (N-1) for '2960', 'CISCO19',' C3750X', 'C4948', 'C891'###################################
###############################################################################################
print ('\nChecking Conditional Flexible Conformance (N-1) for 2960, CISCO19,CISCO29,CISCO39, C3750, C4948, C4900, C891')
conditional_ios = ['12.2(55)','12.2(55)','15.0(2)','15.4(3)','15.4(3)','15.4(3)','16.5.1b','6.0(2)','15.0(2)','15.0(2)SG','16.6.2','15.3(3)','7.0(3)I7',]
conditional_platform = ['2960','2960S','2960X','CISCO19','CISCO29','CISCO39','ISR4','N3K','3750X','C4948','C9300','C891','N9K',]

row_sw = mastersheet.max_row
for i in range(3,row_sw+1):
    for j in range(0,len(conditional_ios)):
        if ((conditional_platform[j] in mastersheet.cell(i,4).value) and (conditional_ios[j] in mastersheet.cell(i,9).value) and (mastersheet.cell(i,9).value != None)):
            if mastersheet.cell(i,12).value != 'EoSWM' and mastersheet.cell(i,14).value != 'EoSWM' and mastersheet.cell(i,16).value != 'LDoS' and mastersheet.cell(i,18).value != 'LDoS':
                if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance'):
                    mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N-1)'
#
#row_sw = mastersheet.max_row
#for i in range(3,row_sw+1):
#    listofplatform = ['2960', 'CISCO29','CISCO19','CISCO39','C3750', 'C4948', 'C4900', 'C891']
#
#    for j in range (0, len(listofplatform)):
#        if listofplatform[j] in mastersheet.cell(i,4).value and mastersheet.cell(i,6).value != 'None' and mastersheet.cell(i,8).value != 'None' and mastersheet.cell(i,8).value != 'null':
#            ############################
#            N = mastersheet.cell(i,6).value
#            N_split = N.split(')')
#            N_1st = N_split[0]
#            N_1st = N_1st.split('(')[0]
#            #print N_1st,1
#            N_2nd = N_split[1]
#            check_last_digit = []
#            for k in range(0, len(N_2nd)):
#                if N_2nd[k] in '0123456789':
#                    check_last_digit.append(N_2nd[k])
#            N_lastdigit = ""
#            N_lastdigit = N_lastdigit.join(check_last_digit)
#            N_lastchar = N_2nd.split(N_lastdigit)[0]
#            N_lastdigit = int(N_lastdigit)
#            ##############################
#            N_1 = mastersheet.cell(i,8).value
#            N_1_split = N_1.split(')')
#            N_1_1st = N_1_split[0]
#            N_1_1st = N_1_1st.split('(')[0]
#            #print N_1_1st,2
#            N_1_2nd = N_1_split[1]
#            check_last_digit = []
#            for k in range(0, len(N_1_2nd)):
#                if N_1_2nd[k] in '0123456789':
#                    check_last_digit.append(N_1_2nd[k])
#            N_1_lastdigit = ""
#            N_1_lastdigit = N_1_lastdigit.join(check_last_digit)
#            N_1_lastchar = N_1_2nd.split(N_1_lastdigit)[0]
#            N_1_lastdigit = int(N_1_lastdigit)
#            ################################
#            N_prod = mastersheet.cell(i,9).value
#            N_prod_split = N_prod.split(')')
#            N_prod_1st = N_prod_split[0]
#            N_prod_1st = N_prod_1st.split('(')[0]
#            #print N_prod_1st,3
#            N_prod_2nd = N_prod_split[1]
#            check_last_digit = []
#            for k in range(0, len(N_prod_2nd)):
#                if N_prod_2nd[k] in '0123456789':
#                    check_last_digit.append(N_prod_2nd[k])
#            N_prod_lastdigit = ""
#            N_prod_lastdigit = N_prod_lastdigit.join(check_last_digit)
#            if N_prod_lastdigit != '':
#                N_prod_lastchar = N_prod_2nd.split(N_prod_lastdigit)[0]
#                N_prod_lastdigit = int(N_prod_lastdigit)
#            ##################################
#        #if '2960' in mastersheet.cell(i,4).value and '2960X' not in mastersheet.cell(i,4).value:
#            if mastersheet.cell(i,12).value != 'EoSWM' and mastersheet.cell(i,14).value != 'EoSWM' and mastersheet.cell(i,16).value != 'LDoS' and mastersheet.cell(i,18).value != 'LDoS':
#                if N_prod_1st == N_1_1st and N_prod_lastchar == N_1_lastchar:
##                    print N_prod_1st
##                    print N_1_1st
##                    print N_prod_lastchar
##                    print N_1_lastchar
#                    
#                    
#                    if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance'):
#                        mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N-1)'
#                elif N_prod_1st == N_1st and N_prod_lastchar == N_lastchar:
##                    print N_prod_1st
##                    print N_1_1st
##                    print N_prod_lastchar
##                    print N_1_lastchar
#                    
#                    if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance' and mastersheet.cell(i,10).value !='Higher than N version in the same release train'):
#                        #print mastersheet.cell(i,10).value
#                        mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N)'
#
#############################################################################################
#####################Conditional Flexible Conformance (N-1) for 'N3K'###################################
################################################################################################
#print '\nChecking Conditional Flexible Conformance (N-1) for N3K'
#row_sw = mastersheet.max_row
#for i in range(3,row_sw+1):
#    listofplatform = ['N3K']
#
#    for j in range (0, len(listofplatform)):
#        if listofplatform[j] in mastersheet.cell(i,4).value and mastersheet.cell(i,6).value != 'None' and mastersheet.cell(i,8).value != 'None' and mastersheet.cell(i,8).value != 'null':
#            ############################
#            N = mastersheet.cell(i,6).value
#            N_split = N.split(')')
#            N_1st = N_split[0]
#            N_1st = N_1st.split('(')[0]
#            N_2nd = N_split[1]
#            N_2nd = list(N_2nd)
#            N_2nd = N_2nd[0]
#            ##############################
#            N_1 = mastersheet.cell(i,8).value
#            N_1_split = N_1.split(')')
#            N_1_1st = N_1_split[0]
#            N_1_1st = N_1_1st.split('(')[0]
#            N_1_2nd = N_1_split[1]
#            N_1_2nd = list(N_1_2nd)
#            N_1_2nd = N_1_2nd[0]
#            ################################
#            N_prod = mastersheet.cell(i,9).value
#            N_prod_split = N_prod.split(')')
#            N_prod_1st = N_prod_split[0]
#            N_prod_1st = N_prod_1st.split('(')[0]
#            N_prod_2nd = N_prod_split[1]
#            N_prod_2nd = list(N_prod_2nd)
#            N_prod_2nd = N_prod_2nd[0]
#            ##################################
#        #if '2960' in mastersheet.cell(i,4).value and '2960X' not in mastersheet.cell(i,4).value:
#            if mastersheet.cell(i,12).value != 'EoSWM' and mastersheet.cell(i,14).value != 'EoSWM' and mastersheet.cell(i,16).value != 'LDoS' and mastersheet.cell(i,18).value != 'LDoS':
#                if N_prod_1st == N_1_1st and N_prod_2nd == N_1_2nd:
#                    if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance'):
#                        mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N-1)'
#                elif N_prod_1st == N_1st and N_prod_2nd == N_1_2nd:
#                    if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance' and mastersheet.cell(i,10).value !='Higher than N version in the same release train'):
#                        mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N)'
#            
#
#############################################################################################
#####################Conditional Flexible Conformance (N-1) for 'ISR4K'###################################
################################################################################################
#print '\nChecking Conditional Flexible Conformance (N-1) for ISR4K'
#row_sw = mastersheet.max_row
#for i in range(3,row_sw+1):
#    listofplatform = ['ISR4']
#
#    for j in range (0, len(listofplatform)):
#        if listofplatform[j] in mastersheet.cell(i,4).value and mastersheet.cell(i,6).value != 'None' and mastersheet.cell(i,8).value != 'None' and mastersheet.cell(i,8).value != 'null':
#            ############################
#            N = mastersheet.cell(i,6).value
#            N_split = N.split('.')
#            N_1st = N_split[0]
#            N_2nd = N_split[1]
#            N_3rd = N_split[2]
#            ##############################
#            N_1 = mastersheet.cell(i,8).value
#            N_1_split = N_1.split('.')
#            N_1_1st = N_1_split[0]
#            N_1_2nd = N_1_split[1]
#            N_1_3rd = N_1_split[2]
#            ################################
#            N_prod = mastersheet.cell(i,9).value
#            N_prod_split = N_prod.split('.')
#            N_prod_1st = N_prod_split[0]
#            N_prod_2nd = N_prod_split[1]
#            N_prod_3rd = N_prod_split[2]
#            ##################################
#        #if '2960' in mastersheet.cell(i,4).value and '2960X' not in mastersheet.cell(i,4).value:
#            if mastersheet.cell(i,12).value != 'EoSWM' and mastersheet.cell(i,14).value != 'EoSWM' and mastersheet.cell(i,16).value != 'LDoS' and mastersheet.cell(i,18).value != 'LDoS':
#                if str(N_prod_1st) == str(N_1_1st) and str(N_prod_2nd) == str(N_1_2nd):
#                    if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance'):
#                        mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N-1)'
#                elif N_prod_1st == N_1st and N_prod_2nd == N_2nd and N_prod_3rd > N_3rd:
#                    if (mastersheet.cell(i,10).value != 'Absolute Conformance' and mastersheet.cell(i,10).value != 'Flexible Conformance' and mastersheet.cell(i,10).value !='Higher than N version in the same release train'):
#                        mastersheet.cell(i,10).value = 'Conditional Flexible Conformance (N)'
#            
############################################################################################
###############Update the NBCU REFRESH REPORT###############################################
#print "\nLoading and Updating the NBCU Refresh Tab...\n\n"
#NBCU_Steve_Update = load_workbook(path_to_save + 'NBCU_Steve_Update.xlsx')
#nbcu_steve_update_filename = NBCU_Steve_Update.get_sheet_by_name('Steve_Update')
#row_update = nbcu_steve_update_filename.max_row
#row_sw = mastersheet.max_row
#for i in range(3,row_sw+1):
#    for j in range(1,row_update+1):
#        if mastersheet.cell(i,2).value == nbcu_steve_update_filename.cell(j,1).value:
#            mastersheet.cell(i,21).value = nbcu_steve_update_filename.cell(j,2).value
#
#NBCU_Steve_Update = ''
#nbcu_steve_update_filename = ''
#

########################################################################################
####update working and 4705bug in cmcs file
#print("\n\n Updating cmcs file and 4705Bug \n\n")
#cmcs_feedback_sheet = mastersheets[2]
#working_sheet = mastersheets[4]
#bug_sheet = mastersheets[5]
#
#row_cmcs = cmcs_feedback_sheet.max_row
#row_working_sheet = working_sheet.max_row
#row_bug_sheet = bug_sheet.max_row
#
#for i in range(3,row_cmcs+1):
#    for j in range(1,row_working_sheet):
#        if str(cmcs_feedback_sheet.cell(i,2).value) == str(working_sheet.cell(j,1).value):
#            cmcs_feedback_sheet.cell(i,26).value = working_sheet.cell(j,2).value
#    for j in range(1,row_bug_sheet+1):
#        if str(cmcs_feedback_sheet.cell(i,2).value) == str(bug_sheet.cell(j,1).value):
#            cmcs_feedback_sheet.cell(i,29).value = bug_sheet.cell(j,2).value
#        
#            
##########################################################################################
########copy NBCU Planning from last week to present week####
print("\n\n Copy NBCU Planning from last week to present week \n\n")
cmcs_feedback_sheet = mastersheets[2]
row_cmcs = cmcs_feedback_sheet.max_row
row_mastersheet = mastersheet.max_row

for i in range(3,row_mastersheet+1):
    for j in range(3,row_cmcs+1):
        if str(mastersheet.cell(i,2).value) == str(cmcs_feedback_sheet.cell(j,2).value):
            mastersheet.cell(i,1).value = cmcs_feedback_sheet.cell(j,1).value
            for k in range(21,32):
                mastersheet.cell(i,k).value = cmcs_feedback_sheet.cell(j,k).value

########################New List from Present Week and mising list from last week##############
print("\n\n New list from present week and Missing list from Last week \n\n")   
cmcs_feedback_sheet = mastersheets[2]            
row_cmcs = cmcs_feedback_sheet.max_row
row_mastersheet = mastersheet.max_row
col_mastersheet = mastersheet.max_column
col_cmcs = cmcs_feedback_sheet.max_column

cmcs_feedback_date = cmcs_feedback_sheet.title
cmcs_feedback_date = cmcs_feedback_date.split('_')
cmcs_feedback_date = cmcs_feedback_date[2]

newlistinpresentweek = 'NewList in ' + time.strftime("%m-%d-%y")
missinglist_fromlastweek = 'Missing from' + cmcs_feedback_date
MasterReport.create_sheet(newlistinpresentweek)
MasterReport.create_sheet(missinglist_fromlastweek)
newlist = MasterReport.get_sheet_by_name(newlistinpresentweek)
missinglist = MasterReport.get_sheet_by_name(missinglist_fromlastweek)

new_device_count_row = 2
newlist_check = 'true'
for i in range(3,row_mastersheet+1):
    newlist_check = 'true'
    for k in range(3,row_cmcs+1):
        if str(mastersheet.cell(i,2).value) == str(cmcs_feedback_sheet.cell(k,2).value):
            newlist_check = 'false'
            break
#        else:
#            newlist_check = 'true'
    if newlist_check == 'true':
        newlist.cell(new_device_count_row,1).value = mastersheet.cell(i,2).value
        new_device_count_row = new_device_count_row+1


missinglist_check = 'true'
missinbg_device_count_row = 2
for i in range(3,row_cmcs+1):
    missinglist_check = 'true'
    for k in range(3,row_mastersheet+1):
        if cmcs_feedback_sheet.cell(i,2).value == mastersheet.cell(k,2).value:
            missinglist_check = 'false'
            break
#        else:
#            missinglist_check = 'true'
    if missinglist_check == 'true':
        missinglist.cell(missinbg_device_count_row,1).value = cmcs_feedback_sheet.cell(i,2).value
        missinbg_device_count_row = missinbg_device_count_row+1

        


newlist_row = newlist.max_row
missinglist_row = missinglist.max_row
newlist.title = newlist.title + '_(' + str(newlist_row-1) + ')'
missinglist.title = missinglist.title + '_(' + str(missinglist_row-1) +')'



####################################################

lastweekreport_name = 'NBCU_SW_HW_LDOS_' + (date_lastreport) + '.xlsx'
lastweekreport_name = path_to_save + lastweekreport_name
lastweekreport_book = load_workbook(lastweekreport_name)
lastweekreport_sheets = lastweekreport_book.worksheets
lastweekreport_sheet = lastweekreport_sheets[2]


cmcs_feedback_sheet = mastersheets[2]
row_mastersheet = mastersheet.max_row
row_cmcs_feedback_sheet = cmcs_feedback_sheet.max_row         
for i in range(3,row_mastersheet+1):
    for j in range(3,row_cmcs_feedback_sheet+1):
        if mastersheet.cell(i,2).value == cmcs_feedback_sheet.cell(j,2).value:
            if mastersheet.cell(i,10) == 'Absolute Conformance' and cmcs_feedback_sheet.cell(j,10).value != 'Absolute Conformance':
                mastersheet.cell(i,32).value = 'YES'
#############################################################
#################################################################

########updating Document Version Control##############################################
print("\n\n Updating Document Version Control from the inventory report recently saved in your folder \n\n")
IT_Inventory_file = NBCU_Inventory_Path_to_Date + '/'+'NBCU_IT_reachable_unreachable_count_'+ time.strftime("%m-%d-%y")+'.xlsx'
BC_Inventory_file = NBCU_Inventory_Path_to_Date + '/'+'Broadcast_reachable_unreachable_count_'+ time.strftime("%m-%d-%y")+'.xlsx'
IT_Inventory_book = load_workbook(IT_Inventory_file)
BC_Inventory_book = load_workbook(BC_Inventory_file)

IT_sheets = IT_Inventory_book.worksheets
BC_sheets = BC_Inventory_book.worksheets
#print IT_sheets

Corp_IT_Total = IT_sheets[1].title
Corp_IT_Total = Corp_IT_Total.split('-')
print (Corp_IT_Total)
Corp_IT_Total = Corp_IT_Total[2]

#Corp_Intl_Total = IT_sheets[2].title
#Corp_Intl_Total = Corp_Intl_Total.split('-')
#Corp_Intl_Total = Corp_Intl_Total[2]
#

Corp_IT_reach = IT_sheets[2].title
Corp_IT_reach = Corp_IT_reach.split('-')
Corp_IT_reach = Corp_IT_reach[1]

#Corp_Intl_reach = IT_sheets[4].title
#Corp_Intl_reach = Corp_Intl_reach.split('-')
#Corp_Intl_reach = Corp_Intl_reach[1]

Corp_IT_unreach = IT_sheets[3].title
Corp_IT_unreach = Corp_IT_unreach.split('-')
Corp_IT_unreach = Corp_IT_unreach[1]

#Corp_Intl_unreach = IT_sheets[6].title
#Corp_Intl_unreach = Corp_Intl_unreach.split('-')
#Corp_Intl_unreach = Corp_Intl_unreach[1]

Corp_IT_confFailed = IT_sheets[4].title
Corp_IT_confFailed = Corp_IT_confFailed.split('-')
Corp_IT_confFailed = Corp_IT_confFailed[1]

#Corp_Intl_confFailed = IT_sheets[8].title
#Corp_Intl_confFailed = Corp_Intl_confFailed.split('-')
#Corp_Intl_confFailed = Corp_Intl_confFailed[1]

BC_Total = BC_sheets[1].title
BC_Total = BC_Total.split('-')
BC_Total = BC_Total[1]

BC_reach = BC_sheets[2].title
BC_reach = BC_reach.split('-')
BC_reach = BC_reach[1]

BC_unreach = BC_sheets[3].title
BC_unreach = BC_unreach.split('-')
BC_unreach = BC_unreach[1]

BC_confFailed = BC_sheets[4].title
BC_confFailed = BC_confFailed.split('-')
BC_confFailed = BC_confFailed[1]

version_control = mastersheets[0]
version_control.cell(6,2).value = int(Corp_IT_Total)
#version_control.cell(7,2).value = int(Corp_Intl_Total)
version_control.cell(8,2).value = int(BC_Total)

version_control.cell(10,2).value = int(Corp_IT_reach)
#version_control.cell(11,2).value = int(Corp_Intl_reach)
version_control.cell(12,2).value = int(BC_reach)

version_control.cell(14,2).value = int(Corp_IT_confFailed)
#version_control.cell(15,2).value = int(Corp_Intl_confFailed)
version_control.cell(16,2).value = int(BC_confFailed)

version_control.cell(18,2).value = int(Corp_IT_unreach)
#version_control.cell(19,2).value = int(Corp_Intl_unreach)
version_control.cell(20,2).value = int(BC_unreach)

#
#######################################################################################
###############################check config collected or not#########################
print("\n\n Checking config collected or Config Failed\n\n")
row_mastersheet = mastersheet.max_row

Corp_IT_confFaileds = IT_sheets[4]
#print Corp_IT_confFaileds.title
#Corp_Intl_confFaileds = IT_sheets[8]
#print Corp_Intl_confFaileds.title
BC_confFaileds = BC_sheets[3]
#print BC_confFaileds.title

US_Config_failed_row = Corp_IT_confFaileds.max_row
#Intl_config_failed_row = Corp_Intl_confFaileds.max_row
BC_confFailed_row = BC_confFaileds.max_row

for i in range(3,row_mastersheet+1):
    configpass = 'true'
    for j in range(2,US_Config_failed_row+1):
        if str(mastersheet.cell(i,2).value) == str(Corp_IT_confFaileds.cell(j,1).value):
            configpass = 'false'
            mastersheet.cell(i,11).value = 'NO'
            #print mastersheet.cell(i,11).value
            break
    if configpass == 'true' and mastersheet.cell(i,11).value != 'NO':
        mastersheet.cell(i,11).value = 'YES'

#for i in range(3,row_mastersheet+1):
#    configpass = 'true'
#    for j in range(2,Intl_config_failed_row+1):
#        if str(mastersheet.cell(i,2).value) == str(Corp_Intl_confFaileds.cell(j,1).value):
#            configpass = 'false'
#            mastersheet.cell(i,11).value = 'NO'
#            #print mastersheet.cell(i,11).value
#            break
#    if configpass == 'true' and mastersheet.cell(i,11).value != 'NO':
#         mastersheet.cell(i,11).value = 'YES'

for i in range(3,row_mastersheet+1):
    configpass = 'true'
    for j in range(2,BC_confFailed_row+1):
        if str(mastersheet.cell(i,2).value) == str(BC_confFaileds.cell(j,1).value):
            configpass = 'false'
            mastersheet.cell(i,11).value = 'NO'
            break
    if configpass == 'true' and mastersheet.cell(i,11).value != 'NO':
        mastersheet.cell(i,11).value = 'YES'

###################################################################################################
                    ############Features and Platform########################
platform = ['FPR','SM-X-ES3','VG','AIR-CT','AIR-WLC','AIR-LAP','AIR-CAP','APIC-SERVER','ASA','CP-7937G',
            'ASR1','ASR-9','C891','CISCO18','CISCO26','CISCO28','CISCO38','CISCO19','CISCO29',
            'CO39','ISR4','IE-4010-4S24P','WS-C2960X','WS-C29','WS-C35','WS-C3750X','WS-C37',
            'WS-C45','WS-C4900M','WS-C49','C65','WS-C36','WS-C38','N2K','N3K','N5K','N6K','N7K','N77','N9K','C94','C93',]

producttype = ['Cisco FirewPower','Switch','Cisco Voice','Wireless','Wireless','Wireless','Wireless',
               'ACI Controller','Firewall','IP Phone','Router','Router','Router','Router',
               'Router','Router','Router','Router','Router','Router','Router','Switch ','Switch',
               'Switch','Switch','Switch','Switch','Switch','Switch','Switch','Switch','Switch',
               'Switch','Switch','Switch','Switch','Switch','Switch','Switch','Switch','Switch','Switch',]

productcatagory = ['Cisco FirewPower','Cisco Ethernet Switch','Voice Gateway','WLC','WLC','Wireless Access Point','Wireless Access Point',
                   'APIC','Cisco ASA','Phone','ASR 1000 Series Router','ASR 9000 Series','Catalyst C800','ISR-G1',
                   'ISR-G1','ISR-G1','ISR-G1','ISR-G2','ISR-G2','ISR-G2','ISR - G3','Industrial Ethernet Switch',
                   'Fixed Switches 2960-X & XR','Switches 2900 series','Switches 3500 series','Fixed Switch 3750-X',
                   'Switches 3700 series','Catalyst 4500','Catalyst 4900 E & M','Catalyst 4900','Catalyst 6500',
                   'Catalyst 3850 / 3650','Catalyst 3850 / 3650','Nexus 2K','Nexus 3K','Nexus 5K','Nexus 6K','Nexus 7K','Nexus 77K','Nexus 9K ACI/Nexus 9K Standalone','Catalyst 9400','Catalyst 9300',]

#print len(platform)
#print len(producttype)
#print len(productcatagory)

row = mastersheet.max_row
mastersheet.insert_cols(5)
mastersheet.insert_cols(6)
mastersheet.cell(2,5).value = 'Product Type'
mastersheet.cell(2,6).value = 'Product Catagory'
for i in range(3,row+1):
    for j in range(0,len(platform)):
        if (platform[j] in mastersheet.cell(i,4).value and mastersheet.cell(i,5).value == None):
            mastersheet.cell(i,5).value = producttype[j]
            mastersheet.cell(i,6).value = productcatagory[j]
            #print mastersheet.cell(i,4).value
            break
#############################################################################################
#############################SnapShot of the Inventory########################################
#snapshot_sheet_name = 'Snapshot of Inventory_' + time.strftime("%m-%d-%y")
#MasterReport.create_sheet(snapshot_sheet_name,2)
#snapshot_sheet = MasterReport.get_sheet_by_name(snapshot_sheet_name)
#
#snapshot_sheet.cell(1,1).value = 'Total Network Element (NE)'
#snapshot_sheet.cell(3,1).value = 'Total Switch'
#snapshot_sheet.cell(5,1).value = 'Platform'
#snapshot_sheet.cell(6,1).value = 'Network Element Count'
#snapshot_sheet.cell(7,1).value = '% Against Total Switches'
#snapshot_sheet.cell(8,1).value = 'Conformance(N, N-1 & N+1)'
#snapshot_sheet.cell(9,1).value = 'Conformance %'
#snapshot_sheet.cell(13,1).value = 'Total Router'
#snapshot_sheet.cell(15,1).value = 'Platform'
#snapshot_sheet.cell(16,1).value = 'Network Element Count'
#snapshot_sheet.cell(17,1).value = '% Against Total Routers'
#snapshot_sheet.cell(18,1).value = 'Conformance(N, N-1 & N+1)'
#snapshot_sheet.cell(19,1).value = 'Conformance %'



##################################################################################################
####################################################################################################
#####################checking N version of the image in last column#############
#####################################################################
for i in range(3, row_mastersheet+1):
    for j in range(3,row_cmcs+1):
        if mastersheet.cell(i,2).value == cmcs_feedback_sheet.cell(j,2).value:
            if str(mastersheet.cell(i,8).value) != str(cmcs_feedback_sheet.cell(j,8).value):
                mastersheet.cell(i,35).value = 'yes'
            break
#


#####################checking update from last week#############
#####################################################################
for i in range(3, row_mastersheet+1):
    for j in range(3,row_cmcs+1):
        if mastersheet.cell(i,2).value == cmcs_feedback_sheet.cell(j,2).value:
            if mastersheet.cell(i,12).value != cmcs_feedback_sheet.cell(j,12).value:
                mastersheet.cell(i,34).value = 'yes'
            break


###########################################################################################################
#########################HW/Chassis EoX Milestone: CY2020####################################################
print("\n Segregating EOX milestone from 19 to 20...")
row = mastersheet.max_row
mastersheet.insert_cols(22)
mastersheet.insert_cols(23)
mastersheet.cell(2,22).value = "HW/Chassis EoX Milestone: CY2020"
mastersheet.cell(2,23).value = "HW/Chassis EoX Milestone Date: CY2020"

for i in range(3,row+1):
#    print (mastersheet.cell(i,21).value)
    if mastersheet.cell(i,21).value is not None:
#        print(i)
        if ("-20" in mastersheet.cell(i,21).value or "-21" in mastersheet.cell(i,21).value or "-22" in mastersheet.cell(i,21).value or "-23" in mastersheet.cell(i,21).value):
#            print(10000)
#            print (i+1)
            mastersheet.cell(i,22).value = mastersheet.cell(i,20).value
            mastersheet.cell(i,23).value = mastersheet.cell(i,21).value
#            print("Inside")

############################################################################################################
######################################hosts for code 16.6.4################################################
limitedhost_16_6_4 = ['fre-knsotv-as01','sac-kcsotv-as01','sac-kcsotv-as02','sac-kcsotv-as03','slc-ktmwtv-as01',
'slc-ktmwtv-as02','ucs-128016-as01','ucs-1280ll-as01','ucs-1280ll-as02','ucs-216008-as01',
'ucs-216008-as02','ucs-345001-as01','ucs-612801-as01','ucs-612802-as01','ucs-612803-as01',
'ucs-612804-as01','ucs-712001-as01','ucs-712002-as01','ucs-712003-as01','ucs-712004-as01',
'ucs-816601-as02','ucs-912801-as03','ucs-sstg24-as01','ucs-sstg25-as01','ucs-sstg26-as01',
'ucs-sstg27-as01',]

row = mastersheet.max_row
for i in range(3,row+1):
    for j in range(0,len(limitedhost_16_6_4)):
        if mastersheet.cell(i,2).value == limitedhost_16_6_4[j]:
            mastersheet.cell(i,8).value = '16.6.6'
            mastersheet.cell(i,9).value = '09 Jul 2019'
            mastersheet.cell(i,10).value = '16.6.4'


###################################################################################################################
##################################################################################################################

#########Save the Docs#############
mastersheet.title = mastersheet.title + '_' + str(int(row_mastersheet)-2)
filename = path_to_save+ 'NBCU_SW_HW_LDOS_' + time.strftime("%m-%d-%y") + '.xlsx'
MasterReport.save(filename)


