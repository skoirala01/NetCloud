import os
import shutil
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
import win32com.client



def nbcu_hw_eox(path_to_read,path_to_save):
####creating new file for LDOS
    HW_EOX = Workbook()
    HW_EOX.create_sheet("NBCU_HW_LDOS_Reached_and_1yr",0)
    hw_eox = HW_EOX.get_sheet_by_name('NBCU_HW_LDOS_Reached_and_1yr')

####reading data    
    filename = path_to_read#'C:/Users/skoirala/Downloads/HW_EOX_DETAILS_20181115_1332.xlsx'
    LDOS = load_workbook(filename)
    ldos_reached = LDOS.get_sheet_by_name('Chassis LDoS Reached')
    ldos_1_yr = LDOS.get_sheet_by_name('Chassis LDoS 1 yr')
    LDoS_1_2_yr = LDOS.get_sheet_by_name('Chassis LDoS 1-2 yr')
    
    
    

####first row heading and bold it
    row = ldos_reached.max_row
    col = ldos_reached.max_column
    for j in range(1, col+1):
        hw_eox.cell(1,j).value = ldos_reached.cell(5,j).value
        hw_eox.cell(1,j).font = Font(sz = 12, b = True)
###data copying  
    a_row = 2
    for i in range (6,row+1):
        for j in range (1,col+1):
            hw_eox.cell(a_row,j).value = ldos_reached.cell(i,j).value        
        a_row = a_row + 1
    
    row = ldos_1_yr.max_row
    col = ldos_1_yr.max_column
    for i in range (6,row+1):
        for j in range (1,col+1):
            hw_eox.cell(a_row,j).value = ldos_1_yr.cell(i,j).value
        a_row = a_row + 1
        
    row = LDoS_1_2_yr.max_row
    col = LDoS_1_2_yr.max_column
    for i in range (6,row+1):
        for j in range (1,col+1):
            hw_eox.cell(a_row,j).value = LDoS_1_2_yr.cell(i,j).value
        a_row = a_row + 1
        
#####sorting    
    row = hw_eox.max_row
    col = hw_eox.max_column
    for i in range(2,row+1):
        for j in range(3,row+1):
            if str(hw_eox.cell(j-1,3).value) > str(hw_eox.cell(j,3).value):
                for k in range(1,col+1):
                    temp = hw_eox.cell(j-1,k).value
                    hw_eox.cell(j-1,k).value = hw_eox.cell(j,k).value
                    hw_eox.cell(j,k).value = temp

####saving
    hw_eox.title = hw_eox.title + '_' + str(row)
    newfilename = path_to_save + 'NBCU_HW_Planning_LDOS_' + time.strftime("%m-%d-%y") + '.xlsx'
    HW_EOX.save(newfilename)
    print ("\n\n\n\n EOX Report 'NBCU_HW_Planning' is Saved, Total = " + str(row) + "\n\n\n")
